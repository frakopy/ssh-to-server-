import os,time, paramiko, datetime,re
from sshtunnel import SSHTunnelForwarder
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.encoders import encode_base64
import smtplib
import threading
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
from rutinas_alarmas_mgw import mgws
from rutinas_alarmas_bsp import llamar_funciones_bsp
from rutinas_alarmas_SIS import llamar_funcion_sis

#---------------------------------------------------------------------------------------------------
#Antes de ejcutar cualquier linea de codigo primero validamos la conexion a internet

from check_internet import revisar_internet, reconectar_wifi

internet = revisar_internet()
        
if internet == True:
    print('Conexión a internet exitosa!!!\n')

elif internet == False:
    reconectar_wifi()

#----------------------REALIZAR CONEXION A LA VPN CISCO-----------------------------------------------------------------------------

os.chdir('C:\Program Files (x86)\Cisco Systems\VPN Client')
os.system('vpnclient.exe connect vpn-switch user xxxx pwd xxxxxx')

#-----------------------------------------------------------------------------------------------------------------------------------
#Obtenemos la fecha de ayer y hoy para abir los archivos de ayer y hoy, leer su informacion y
#compararlos para ver si existen alarmas nuevas
fecha_hoy = datetime.datetime.now()
fecha_ayer = fecha_hoy - datetime.timedelta(days=1)

fecha_hoy = str(fecha_hoy)
fecha_ayer = str(fecha_ayer)

#-------------------------Definicion de las funciones-------------------------------------------------------------------------
def enviar_correo():

    archivo = 'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/Resumen_Alarmas/Resumen_Alarmas_MSCs_MGWs.xlsx'
    msg = MIMEMultipart('mixed')
    mensaje = 'Hola Ingeniero,\n\nAdjunto el reporte de alarmas.\n\nSaludos.'

    msg['From'] = 'xxxxxxxxx@gmail.com'
    msg['To'] = 'xxxxxxxxxx@millicom.com'
    msg['Subject'] = 'Reporte_Alarmas_MSCs_MGWs'

    msg.attach(MIMEText(mensaje, 'plain'))

    attachment = open(archivo, 'rb')

    p = MIMEBase('application', "vnd.ms-excel")
    p.set_payload((attachment).read())
    encode_base64(p)
    p.add_header('Content-Disposition',"attachment; filename= %s" % 'Resumen_Alarmas_MSCs_MGWs.xlsx')
    msg.attach(p)

    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(msg['From'], 'xxxxxxx')
    server.sendmail(msg['From'], msg['To'], msg.as_string())
    server.quit()

def obtener_alarmas(datos,puerto,node_name):
    
    try:
        with SSHTunnelForwarder(
            ('ip', port),
            ssh_username='xxxxxx',
            ssh_password='xxxxxx',
            remote_bind_address=(datos[0], 22),
            local_bind_address=('0.0.0.0', puerto)
            ) as tunnel:
                ssh = paramiko.SSHClient()
                ssh.load_system_host_keys()
                ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                ssh.connect('127.0.0.1',puerto, username=datos[1], password=datos[2]) 
                
                salida = ''
                ruta = 'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/'+node_name+'/'
                
                if node_name == 'NICMA2M':
                    stdin, stdout, stderr = ssh.exec_command('mml')
                    stdin.write('allip;')
                    stdin.flush()
                    stdin.close()
                    salida = stdout.readlines()
                    
                    with open(ruta+datos[3],'w') as f:
                        for s in salida:
                            f.write(s)

                    ssh.close()
                    resultado = 'Se capturaron los datos con éxito en el nodo: ' + node_name
                    return resultado

                else:
                    t = ssh.get_transport()
                    chan = t.open_session()
                    chan.get_pty()

                    chan = ssh.invoke_shell()
                    comandos = ['mml allip;\n', 'mml -cp cp1 allip;\n','mml -cp cp2 allip;\n']

                    for comando in comandos:
                        chan.send(comando)  
                        time.sleep(3)
                        while not chan.recv_ready(): 
                            time.sleep(3)
                        salida += chan.recv(99999999999999).decode('ASCII')

                    file = open(ruta+datos[3],'w') 
                    file.write(salida)
                    file.close()
                    
                    ssh.close()#Cerramos la conexion ssh para que se puedan hacer otras conexiones

                    resultado = 'Se capturaron los datos con éxito en el nodo: ' + node_name
                    return resultado  

    except:
        os.system('cls')
        print('\n'+' CREDENCIALES INCORRECTAS O PROBLEMAS DE RED '.center(80,'='))

def limpar_datos():
    
    #Cargamos el archivo Excell, obtenemos las hojas que este contiene y las asignamos a distintas varialbles
    excell_path = 'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/Resumen_Alarmas/Resumen_Alarmas_MSCs_MGWs.xlsx'
    archivo_excell = load_workbook(excell_path)

    hoja_nicma4m = archivo_excell['NICMA4M']
    hoja_nicgr1m = archivo_excell['NICGR1M']
    hoja_nicma2m = archivo_excell['NICMA2M']
    hoja_pansl4m = archivo_excell['PANSL4M']
    hoja_davte3m = archivo_excell['DAVTE3M']
    hoja_sjlin1m = archivo_excell['SJLIN1M']
    hoja_nicma1g = archivo_excell['NICMA1G']
    hoja_nicma2g = archivo_excell['NICMA2G']
    hoja_nicma3g = archivo_excell['NICMA3G']
    hoja_nicgr1g = archivo_excell['NICGR1G']
    hoja_davte1g = archivo_excell['DAVTE1G']
    hoja_pansl3g = archivo_excell['PANSL3G']
    hoja_bsp_nicgr1 = archivo_excell['BSP_NICGR1M']
    hoja_bsp_davte3 = archivo_excell['BSP_DAVTE3M']
    hoja_sis_nicma4 = archivo_excell['SIS_NICMA4M']
    hoja_sis_pan4 = archivo_excell['SIS_PANSL4M']


    #Obtnemos la cantidad de filas que contiene cada hoja para saber cuantas filas debemos eliminar
    filas_nicma4m = hoja_nicma4m.max_row
    filas_nicgr1m = hoja_nicgr1m.max_row
    filas_nicma2m = hoja_nicma2m.max_row
    filas_pansl4m = hoja_pansl4m.max_row
    filas_davte3m = hoja_davte3m.max_row
    filas_sjlin1m = hoja_sjlin1m.max_row
    filas_nicma1g = hoja_nicma1g.max_row
    filas_nicma2g = hoja_nicma2g.max_row
    filas_nicma3g = hoja_nicma3g.max_row
    filas_nicgr1g = hoja_nicgr1g.max_row
    filas_davte1g = hoja_davte1g.max_row
    filas_pansl3g = hoja_nicma1g.max_row
    filas_bsp_nicgr1 = hoja_bsp_nicgr1.max_row
    filas_bsp_davte3 = hoja_bsp_davte3.max_row
    filas_sis_nicma4 = hoja_sis_nicma4.max_row
    filas_sis_pan4 = hoja_sis_pan4.max_row


    hojas_filas = [
        (hoja_nicma4m,filas_nicma4m),
        (hoja_nicgr1m,filas_nicgr1m),
        (hoja_nicma2m,filas_nicma2m),
        (hoja_pansl4m,filas_pansl4m),
        (hoja_davte3m,filas_davte3m),
        (hoja_sjlin1m,filas_sjlin1m),
        (hoja_nicma1g,filas_nicma1g),
        (hoja_nicma2g,filas_nicma2g),
        (hoja_nicma3g,filas_nicma3g),
        (hoja_nicgr1g,filas_nicgr1g),
        (hoja_davte1g,filas_davte1g),
        (hoja_pansl3g,filas_pansl3g),
        (hoja_bsp_nicgr1,filas_bsp_nicgr1),
        (hoja_bsp_davte3,filas_bsp_davte3),
        (hoja_sis_nicma4,filas_sis_nicma4),
        (hoja_sis_pan4,filas_sis_pan4)
        ]

    #Eliminamos la informacion que se encuentra en cada celda de cada hoja
    for hoja_nodo, nfilas in hojas_filas:
        for i in range(2, nfilas + 1):
            i = str(i)
            hoja_nodo['A'+i] = ''
            hoja_nodo['B'+i] = ''
            hoja_nodo['C'+i] = ''

    #Guardamos el archivo despues de borrar los datos
    archivo_excell.save(excell_path)


def generar_archivo_excell(id_nodo,nombre_archivo):

    ruta_arhivos_alarmas = 'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/'

    #Agrupamos las alarmas en una lista para iterarlas mas adelante:
    ruta_arhivo = ruta_arhivos_alarmas+id_nodo+'/'+nombre_archivo 
    alarmas = ''       
    with open(ruta_arhivo, 'r') as file:
        for linea in file:
            if '**' in linea or 'PLEASE EXIT' in linea or 'allip;' in linea or 'ALARM LIST' in linea or '>' in linea or 'END' in linea:
                linea = ''
            if re.match('\D\d/\D{3}', linea):
                linea = '*'+linea
            alarmas += linea

    alarmas = alarmas.split('*')

    #Iteramos las alarmas y las escribimos en la hoja correspondiente del archivo Excell:
    ruta_archivo = 'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/Resumen_Alarmas/Resumen_Alarmas_MSCs_MGWs.xlsx'
    excell_file = load_workbook(ruta_archivo)
    nombre_hoja = nombre_archivo[8:15]

    hoja = excell_file[nombre_hoja]

    c = 2
    for alm in alarmas:
        try:
            categoria = re.findall('\D\d/\D{3}', alm)
            fecha = re.findall('\d{6}', alm)
            fecha = fecha[0][0:2]+'-'+fecha[0][2:4]+'-'+fecha[0][4:6]
            hoja['A'+str(c)] = fecha
            hoja['B'+str(c)] = categoria[0]
            hoja['C'+str(c)] = alm
            c +=1 
        except:
            pass

    #Guardamos el archivo despues de borrar los datos
    excell_file.save(ruta_archivo)


def depura_antiguos():
    
    ahora = time.time()

    rutas = [
        'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/DAVTE3M',
        'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/NICGR1M',
        'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/NICMA2M',
        'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/NICMA4M',
        'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/PANSL4M',
        'D:/A_PYTHON/ProgramasPython/Control_NodosCA/Scritps_Rutinas/Rutinas_Alarmas/SJLIN1M'
        ]

    for ruta in rutas:
        for file in os.listdir(ruta):
            if os.stat(ruta+'/'+file).st_mtime < (ahora - 2592000) and file.endswith('.txt'): 
                os.remove(ruta+'/'+file)

    #verificamos cada archivo para saber si tiene fecha de creación mayor de 30 dias y si es asi
    #entonces que lo borre, en caso contrario no hace nada, 2592000 son 30 dias en segundos

#---------------------------------------------------------------------------------------------------

lista_nodos = [
    'ALARMAS_NICMA4M','ALARMAS_NICGR1M','ALARMAS_NICMA2M',
    'ALARMAS_PANSL4M','ALARMAS_DAVTE3M','ALARMAS_SJLIN1M'
            ]

datos = {
    'NICMA4M':['IP','USER','PASSS','ALARMAS_NICMA4M'+'_'+fecha_hoy[:10]+'.txt'],
    'NICGR1M':['IP','USER','PASSS','ALARMAS_NICGR1M'+'_'+fecha_hoy[:10]+'.txt'],
    'NICMA2M':['IP','USER','PASSS','ALARMAS_NICMA2M'+'_'+fecha_hoy[:10]+'.txt'],
    'PANSL4M':['IP','USER','PASSS','ALARMAS_PANSL4M'+'_'+fecha_hoy[:10]+'.txt'],
    'DAVTE3M':['IP','USER','PASSS','ALARMAS_DAVTE3M'+'_'+fecha_hoy[:10]+'.txt'],
    'SJLIN1M':['IP','USER','PASSS','ALARMAS_SJLIN1M'+'_'+fecha_hoy[:10]+'.txt'],
        }


puerto = 65031
ejecutor = ThreadPoolExecutor(max_workers=6)
futuros = []
for l in lista_nodos:
    os.system('cls')
    print('\n1-Se va a iniciar con la busqueda de nuevas alarmas para hoy', fecha_hoy[:-7] + '\n')
    node_name = l[8:]
    futuros.append(ejecutor.submit(obtener_alarmas, datos[node_name], puerto, node_name))
    puerto += 1

#Tomar nota que ejecutor.submit retorna instancia de tipo futuro por lo que cada instancia la vamos
#guardando en una lista que hemos declarado con el nombre futuros


#La funcion as_completed espera a que se complete cada llamada a la funcion obtener_alarmas, por lo
#que la ejecución del programa se detiene en esta linea y continua hasta que cada una de las llamadas
#a la funcion obtener_alarmas se haya completado
for futuro in concurrent.futures.as_completed(futuros):
    print(futuro.result())

#las lineas de arriba son necesarias para evitar que el hilo principal se siga ejecutando mientras
#se ejecutan los hilos hijos ya que necesiamos que primero se obtengan las alarmas para luego poder
#usar esa informacion, otra opcion podria ser usar el metodo join de los threads(hilos) para indicarle
#hilo padre que espera por la finalizacion de sus hijos para continuar con la ejecucion del programa
#---------------------------------------------------------------------------------------------------

#Llamamos a esta funcion para eliminar registros antiguos en el archivo Excell
print('\n2-Ahora se procedera a eliminar los datos antiguos del archivo Excell que sera enviado por correo')
limpar_datos()

#---------------------------------------------------------------------------------------------------

#Mediante un ciclo for vamos llamando a la funcion generar_archivo_excell para que escriba los datos
print('\n3-A continuación se procede con la escritura de los datos sobre el archivo Excell')
for l in lista_nodos:
    id_nodo = l[8:]
    nombre_archivo = l + '_' + fecha_hoy[:10] + '.txt'
    generar_archivo_excell(id_nodo,nombre_archivo)
#---------------------------------------------------------------------------------------------------
#llamamos a la funcion importada que se encuentra en el archivo rutinas_alarmas_mgw.py la cual se 
#encarga de obtener las alarmas de los MGW y luego las escribe en el archivo excell.
mgws()

#---------------------------------------------------------------------------------------------------

#llamamos a la funcion que contrala las llamadas al resto de funciones
llamar_funciones_bsp()

#---------------------------------------------------------------------------------------------------
#llamamos a la funcion que contrala las llamadas al resto de funciones
llamar_funcion_sis()
#---------------------------------------------------------------------------------------------------


#Por ultimo llamamos a la funcion que envia el archivo Excell con el resumen de las alarmas por correo
print('\n4-Enviando correo...')
enviar_correo()

#---------------------------------------------------------------------------------------------------

#Se llama a la funcion que verifica si existen archivos con fecha de creacion mayor a 30 dias
#y si es asi entonces los elimina, caso contrario no hace nada.
print('\n5-Por último se verifica si existen archivos con mas de 30 días de creación para depurarlos')
depura_antiguos()

#---------------------------------------------------------------------------------------------------
